"""Convert .xlsx and .xls spreadsheets to markdown.

Dispatches by file extension: openpyxl for .xlsx, xlrd for .xls.
Uses cached (data_only) values — formulas are not evaluated.
Sheets with more than 500 rows are truncated with a warning note.
"""

from pathlib import Path

_ROW_LIMIT = 500


def _escape(value: object) -> str:
    text = "" if value is None else str(value)
    return text.replace("|", "\\|").replace("\n", " ").replace("\r", "")


def _safe_sheet_name(name: str) -> str:
    for ch in r"\`*_{}[]()#+-.!|":
        name = name.replace(ch, " ")
    return name.strip()


def _rows_to_md(headers: list[str], rows: list[list[str]], truncated: bool) -> str:
    lines: list[str] = []
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    for row in rows:
        lines.append("| " + " | ".join(row) + " |")
    if truncated:
        lines.append(f"\n> ⚠ Truncated at {_ROW_LIMIT} rows.")
    return "\n".join(lines)


def _convert_xlsx(src: Path, dst: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(src, read_only=True, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # Determine max populated column width from header row
        header_raw = all_rows[0]
        col_count = len(header_raw)
        # Trim trailing None columns from header
        while col_count > 1 and header_raw[col_count - 1] is None:
            col_count -= 1

        headers = [_escape(header_raw[i]) for i in range(col_count)]
        body_rows = all_rows[1:]
        truncated = len(body_rows) > _ROW_LIMIT
        body_rows = body_rows[:_ROW_LIMIT]

        md_rows = []
        for row in body_rows:
            cells = [_escape(row[i] if i < len(row) else None) for i in range(col_count)]
            md_rows.append(cells)

        sections.append(f"## {_safe_sheet_name(sheet_name)}\n\n{_rows_to_md(headers, md_rows, truncated)}")

    wb.close()
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text("\n\n".join(sections), encoding="utf-8")


def _convert_xls(src: Path, dst: Path) -> None:
    import xlrd

    wb = xlrd.open_workbook(str(src))
    sections: list[str] = []

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        if ws.nrows == 0:
            continue

        header_raw = ws.row_values(0)
        col_count = len(header_raw)
        while col_count > 1 and header_raw[col_count - 1] in (None, ""):
            col_count -= 1

        headers = [_escape(header_raw[i]) for i in range(col_count)]
        nrows = ws.nrows - 1
        truncated = nrows > _ROW_LIMIT
        nrows = min(nrows, _ROW_LIMIT)

        md_rows = []
        for r in range(1, nrows + 1):
            row = ws.row_values(r)
            cells = [_escape(row[i] if i < len(row) else None) for i in range(col_count)]
            md_rows.append(cells)

        sections.append(f"## {_safe_sheet_name(sheet_name)}\n\n{_rows_to_md(headers, md_rows, truncated)}")

    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text("\n\n".join(sections), encoding="utf-8")


def convert(src: Path, dst: Path) -> None:
    if src.suffix.lower() == ".xls":
        _convert_xls(src, dst)
    else:
        _convert_xlsx(src, dst)
