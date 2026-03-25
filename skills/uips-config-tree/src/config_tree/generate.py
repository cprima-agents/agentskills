"""
generate — Config.xlsx → typed C# classes and XAML clipboard snippet.

Reads a UiPath Config.xlsx and generates:
  - A typed C# file with one class per sheet plus a CodedConfig aggregator
  - Optionally a ClipboardData XAML snippet for paste into UiPath Studio
"""

import datetime
import importlib.resources
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from string import Template

import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_FILE = Path("./Data/Config.xlsx")
DEFAULT_NAMESPACE = "Cpmf.Config"
DEFAULT_ROOT_CLASS = "CodedConfig"
DEFAULT_DOTNET = "net6"
DEFAULT_UIPATH_VAR = "out_ConFigTree"

_HEADER_STANDARD = ["name", "value", "description"]
_HEADER_ASSET = ["name", "asset", "orchestratorassetfolder", "description"]
_SYSTEM_TYPES = {"DateOnly", "DateTime", "TimeOnly"}

# ---------------------------------------------------------------------------
# Intermediate representation
# ---------------------------------------------------------------------------


@dataclass
class PropertyDef:
    prop_name: str
    cs_type: str
    description: str | None
    asset_name: str | None = None
    folder: str | None = None


@dataclass
class SheetDef:
    sheet_name: str
    class_name: str
    schema: str  # "standard" or "asset"
    properties: list[PropertyDef] = field(default_factory=list)


@dataclass
class WorkbookIR:
    root_class: str
    namespace: str
    dotnet_version: str
    emit_xml_docs: bool
    generate_loader: bool
    generate_tostring: bool
    generate_tojson: bool
    generate_pristine: bool
    readonly: bool
    uipath_var: str
    sheets: list[SheetDef] = field(default_factory=list)
    needs_orchestrator_asset: bool = False
    needs_system_using: bool = False


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------


def load_workbook(path: Path) -> openpyxl.Workbook:
    if path.suffix.lower() != ".xlsx":
        print(f"Error: expected a .xlsx file, got: {path}", file=sys.stderr)
        sys.exit(1)
    if not path.exists():
        print(f"Error: file not found: {path}", file=sys.stderr)
        sys.exit(1)
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        print(f"Error: cannot open {path}: {exc}", file=sys.stderr)
        sys.exit(1)


def detect_schema(header_row: list) -> str | None:
    normalised = [str(v).strip().lower() if v is not None else "" for v in header_row]
    if len(normalised) >= 3 and normalised[:3] == _HEADER_STANDARD[:3]:
        return "standard"
    if len(normalised) >= 4 and normalised[:4] == _HEADER_ASSET:
        return "asset"
    return None


def read_sheet(ws) -> tuple[str, list[dict]]:
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows or all(v is None for v in all_rows[0]):
        print(f"  [WARN] Sheet '{ws.title}': no header row found, skipping.", file=sys.stderr)
        return "skip", []

    header = list(all_rows[0])
    schema = detect_schema(header)
    if schema is None:
        print(f"  [WARN] Sheet '{ws.title}': unrecognised header {header[:4]}, skipping.", file=sys.stderr)
        return "skip", []

    rows = []
    for raw in all_rows[1:]:
        if not raw or raw[0] is None or str(raw[0]).strip() == "":
            continue
        if schema == "standard":
            name = str(raw[0]).strip()
            value = raw[1] if len(raw) > 1 else None
            desc = str(raw[2]).strip() if len(raw) > 2 and raw[2] is not None else None
            rows.append({"name": name, "value": value, "description": desc})
        else:
            name = str(raw[0]).strip()
            asset_name = str(raw[1]).strip() if len(raw) > 1 and raw[1] is not None else ""
            folder = str(raw[2]).strip() if len(raw) > 2 and raw[2] is not None else ""
            desc = str(raw[3]).strip() if len(raw) > 3 and raw[3] is not None else None
            rows.append({"name": name, "asset_name": asset_name, "folder": folder, "description": desc})

    return schema, rows


def infer_cs_type(py_value, dotnet_version: str) -> str:
    if py_value is None:
        return "string"
    if isinstance(py_value, bool):
        return "bool"
    if isinstance(py_value, int):
        return "int"
    if isinstance(py_value, float):
        return "double"
    if isinstance(py_value, datetime.time) and not isinstance(py_value, datetime.datetime):
        return "TimeOnly"
    if isinstance(py_value, datetime.datetime):
        if py_value.hour == 0 and py_value.minute == 0 and py_value.second == 0 and py_value.microsecond == 0:
            return "DateOnly"
        return "DateTime"
    if isinstance(py_value, datetime.date):
        return "DateOnly"
    return "string"


def to_pascal_case(name: str) -> str:
    segments = re.split(r"[^a-zA-Z0-9]+", name)
    result = "".join(seg[0].upper() + seg[1:] for seg in segments if seg)
    if result and result[0].isdigit():
        result = "_" + result
    return result


def sheet_name_to_class_name(sheet_name: str) -> str:
    segments = re.split(r"[.\-]", sheet_name)
    result = "".join(seg[0].upper() + seg[1:] for seg in segments if seg)
    return result + "Config"


# ---------------------------------------------------------------------------
# Build IR
# ---------------------------------------------------------------------------


def build_ir(
    wb: openpyxl.Workbook,
    sheet_filter: list[str] | None,
    namespace: str,
    root_class: str,
    dotnet_version: str,
    emit_xml_docs: bool,
    generate_loader: bool,
    generate_tostring: bool,
    generate_tojson: bool,
    generate_pristine: bool,
    readonly: bool,
    uipath_var: str,
) -> WorkbookIR:
    ir = WorkbookIR(
        root_class=root_class,
        namespace=namespace,
        dotnet_version=dotnet_version,
        emit_xml_docs=emit_xml_docs,
        generate_loader=generate_loader,
        generate_tostring=generate_tostring,
        generate_tojson=generate_tojson,
        generate_pristine=generate_pristine,
        readonly=readonly,
        uipath_var=uipath_var,
    )

    requested = set(sheet_filter) if sheet_filter else None
    if requested:
        for name in requested:
            if name not in wb.sheetnames:
                print(f"  [WARN] Requested sheet '{name}' not found in workbook.", file=sys.stderr)

    for ws in wb.worksheets:
        if requested and ws.title not in requested:
            continue

        schema, rows = read_sheet(ws)
        if schema == "skip":
            continue

        class_name = sheet_name_to_class_name(ws.title)
        sheet_def = SheetDef(sheet_name=ws.title, class_name=class_name, schema=schema)

        for row in rows:
            prop_name = to_pascal_case(row["name"])
            if schema == "standard":
                cs_type = infer_cs_type(row["value"], dotnet_version)
                sheet_def.properties.append(
                    PropertyDef(prop_name=prop_name, cs_type=cs_type, description=row["description"])
                )
                if cs_type in _SYSTEM_TYPES:
                    ir.needs_system_using = True
            else:
                sheet_def.properties.append(
                    PropertyDef(
                        prop_name=prop_name,
                        cs_type="OrchestratorAsset",
                        description=row["description"],
                        asset_name=row["asset_name"],
                        folder=row["folder"],
                    )
                )

        if schema == "asset":
            ir.needs_orchestrator_asset = True

        ir.sheets.append(sheet_def)

    return ir


# ---------------------------------------------------------------------------
# C# generation
# ---------------------------------------------------------------------------


def _accessor(readonly: bool) -> str:
    return "{ get; init; }" if readonly else "{ get; set; }"


def _xml_doc(description: str | None, emit: bool, indent: str) -> str:
    if not emit or not description:
        return ""
    safe = description.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return f"{indent}/// <summary>{safe}</summary>\n"


def _property_initializer(cs_type: str) -> str:
    if cs_type == "string":
        return ' = "";'
    if cs_type == "OrchestratorAsset":
        return " = new();"
    return ""


def _from_data_table_case(prop: PropertyDef) -> list[str]:
    name = prop.prop_name
    lines = []
    if prop.cs_type == "OrchestratorAsset":
        lines.append(f'                case "{name}":')
        lines.append(f'                    cfg.{name}.AssetName = row[1]?.ToString()?.Trim() ?? "";')
        lines.append(f'                    cfg.{name}.Folder    = row[2]?.ToString()?.Trim() ?? "";')
        lines.append('                    break;')
    elif prop.cs_type == "string":
        lines.append(f'                case "{name}": cfg.{name} = value; break;')
    elif prop.cs_type in ("int", "double", "bool", "DateTime"):
        t = prop.cs_type
        lines.append(f'                case "{name}":')
        lines.append(f'                    if ({t}.TryParse(value, out var v_{name})) cfg.{name} = v_{name};')
        lines.append('                    break;')
    elif prop.cs_type == "DateOnly":
        lines.append(f'                case "{name}":')
        lines.append(f'                    if (DateOnly.TryParse(value, out var v_{name})) cfg.{name} = v_{name};')
        lines.append(f'                    else if (DateTime.TryParse(value, out var dt_{name})) cfg.{name} = DateOnly.FromDateTime(dt_{name});')
        lines.append('                    break;')
    elif prop.cs_type == "TimeOnly":
        lines.append(f'                case "{name}":')
        lines.append(f'                    if (TimeOnly.TryParse(value, out var v_{name})) cfg.{name} = v_{name};')
        lines.append(f'                    else if (DateTime.TryParse(value, out var dt_{name})) cfg.{name} = TimeOnly.FromDateTime(dt_{name});')
        lines.append('                    break;')
    else:
        lines.append(f'                case "{name}": cfg.{name} = value; break;')
    return lines


def render_cs_orchestrator_asset_class(readonly: bool) -> str:
    acc = _accessor(readonly)
    return (
        f"    public class OrchestratorAsset\n"
        f"    {{\n"
        f'        public string AssetName {acc} = "";\n'
        f'        public string Folder {acc} = "";\n'
        f"    }}"
    )


def render_cs_sheet_class(sheet: SheetDef, ir: WorkbookIR) -> str:
    acc = _accessor(ir.readonly)
    lines = [f"    public class {sheet.class_name}", "    {"]

    for prop in sheet.properties:
        doc = _xml_doc(prop.description, ir.emit_xml_docs, "        ")
        if doc:
            lines.append(doc.rstrip("\n"))
        init = _property_initializer(prop.cs_type)
        lines.append(f"        public {prop.cs_type} {prop.prop_name} {acc}{init}")

    if ir.generate_loader:
        lines.append("")
        lines.append(f"        public static {sheet.class_name} FromDataTable(DataTable dt)")
        lines.append("        {")
        if sheet.properties:
            lines.append(f"            var cfg = new {sheet.class_name}();")
            lines.append("            foreach (DataRow row in dt.Rows)")
            lines.append("            {")
            if sheet.schema == "standard":
                lines.append('                var key   = row[0]?.ToString()?.Trim();')
                lines.append('                var value = row[1]?.ToString()?.Trim() ?? "";')
            else:
                lines.append('                var key = row[0]?.ToString()?.Trim();')
            lines.append("                switch (key)")
            lines.append("                {")
            for prop in sheet.properties:
                lines.extend(_from_data_table_case(prop))
            lines.append("                }")
            lines.append("            }")
            lines.append("            return cfg;")
        else:
            lines.append(f"            return new {sheet.class_name}();")
        lines.append("        }")

    lines.append("    }")
    return "\n".join(lines)


def render_cs_root_class(ir: WorkbookIR) -> str:
    acc = _accessor(ir.readonly)
    lines = []
    if ir.emit_xml_docs:
        lines.append("    /// <summary>Root configuration object.</summary>")
    lines.append(f"    public class {ir.root_class}")
    lines.append("    {")

    for sheet in ir.sheets:
        prop_name = sheet.class_name[:-6]
        lines.append(f"        public {sheet.class_name} {prop_name} {acc} = new();")

    if ir.generate_loader:
        lines.append("")
        lines.append(f"        public static {ir.root_class} Load(Dictionary<string, DataTable> tables)")
        lines.append("        {")
        lines.append(f"            var cfg = new {ir.root_class}();")
        for sheet in ir.sheets:
            prop_name = sheet.class_name[:-6]
            var_name = f"t_{prop_name}"
            lines.append(
                f'            if (tables.TryGetValue("{sheet.sheet_name}", out var {var_name}))'
                f' cfg.{prop_name} = {sheet.class_name}.FromDataTable({var_name});'
            )
        lines.append("            return cfg;")
        lines.append("        }")

    if ir.generate_pristine:
        lines.append("")
        lines.append("        public static readonly Dictionary<string, HashSet<string>> Schema = new()")
        lines.append("        {")
        for sheet in ir.sheets:
            keys = ", ".join(f'"{p.prop_name}"' for p in sheet.properties)
            lines.append(f'            ["{sheet.sheet_name}"] = new HashSet<string> {{ {keys} }},')
        lines.append("        };")
        lines.append("")
        lines.append("        public static DriftReport CheckDrift(Dictionary<string, DataTable> tables)")
        lines.append("        {")
        lines.append("            var report = new DriftReport();")
        lines.append("            foreach (var kvp in Schema)")
        lines.append("            {")
        lines.append('                if (!tables.TryGetValue(kvp.Key, out var dt)) { report.MissingSheets.Add(kvp.Key); continue; }')
        lines.append('                var actualKeys = new HashSet<string>(dt.AsEnumerable().Select(r => r[0]?.ToString()?.Trim() ?? ""));')
        lines.append('                report.UnknownKeys.AddRange(actualKeys.Except(kvp.Value).Select(k => $"{kvp.Key}.{k}"));')
        lines.append('                report.MissingKeys.AddRange(kvp.Value.Except(actualKeys).Select(k => $"{kvp.Key}.{k}"));')
        lines.append("            }")
        lines.append("            return report;")
        lines.append("        }")

    if ir.generate_tostring:
        parts = ", ".join(
            f"{sheet.class_name[:-6]} = {{{sheet.class_name[:-6]}}}"
            for sheet in ir.sheets
        )
        lines.append("")
        lines.append("        public override string ToString() =>")
        lines.append(f'            $"{ir.root_class} {{{{ {parts} }}}}";')

    if ir.generate_tojson:
        lines.append("")
        lines.append("        public string ToJson() =>")
        lines.append("            System.Text.Json.JsonSerializer.Serialize(this);")

    lines.append("    }")
    return "\n".join(lines)


def render_cs_drift_report_class() -> str:
    return (
        "    public class DriftReport\n"
        "    {\n"
        '        public List<string> MissingSheets { get; set; } = new();\n'
        '        public List<string> MissingKeys { get; set; } = new();\n'
        '        public List<string> UnknownKeys { get; set; } = new();\n'
        "        public bool HasDrift => MissingSheets.Any() || MissingKeys.Any() || UnknownKeys.Any();\n"
        "        public override string ToString() => HasDrift\n"
        '            ? $"Drift detected: {MissingSheets.Count} missing sheet(s), {MissingKeys.Count} missing key(s), {UnknownKeys.Count} unknown key(s)"\n'
        '            : "No drift detected";\n'
        "    }"
    )


def render_cs(ir: WorkbookIR) -> str:
    using_lines = []
    if ir.needs_system_using:
        using_lines.append("using System;")
    if ir.generate_loader or ir.generate_pristine:
        using_lines.append("using System.Collections.Generic;")
    if ir.generate_loader:
        using_lines.append("using System.Data;")
    if ir.generate_pristine:
        using_lines.append("using System.Linq;")
    if ir.generate_tojson:
        using_lines.append("using System.Text.Json;")

    usings = "\n".join(using_lines) + "\n\n" if using_lines else ""

    body_parts = [render_cs_root_class(ir)]
    if ir.generate_pristine:
        body_parts.append(render_cs_drift_report_class())
    if ir.needs_orchestrator_asset:
        body_parts.append(render_cs_orchestrator_asset_class(ir.readonly))
    for sheet in ir.sheets:
        body_parts.append(render_cs_sheet_class(sheet, ir))

    body = "\n\n".join(body_parts) + "\n"

    res = importlib.resources.files("config_tree.resources")
    tmpl = Template((res / "CodedConfig.cs.template").read_text(encoding="utf-8"))
    return tmpl.substitute(usings=usings, namespace=ir.namespace, body=body)


# ---------------------------------------------------------------------------
# XAML generation
# ---------------------------------------------------------------------------


def render_xaml(ir: WorkbookIR) -> str:
    var = ir.uipath_var
    root = ir.root_class
    load_label = var.removeprefix("out_") if var.startswith("out_") else var

    res = importlib.resources.files("config_tree.resources")
    tmpl = Template((res / "LoadTypedConfig.xaml.template").read_text(encoding="utf-8"))
    return tmpl.substitute(var=var, root=root, load_label=load_label)


# ---------------------------------------------------------------------------
# Diagnostics
# ---------------------------------------------------------------------------


def emit_diagnostics(ir: WorkbookIR, fmt: str) -> None:
    if fmt == "json":
        data = {
            "sheets": [
                {"name": s.sheet_name, "class": s.class_name, "schema": s.schema, "properties": len(s.properties)}
                for s in ir.sheets
            ],
            "needs_orchestrator_asset": ir.needs_orchestrator_asset,
            "needs_system_using": ir.needs_system_using,
        }
        print(json.dumps(data, indent=2), file=sys.stderr)
    else:
        print(f"ConFormMold: {len(ir.sheets)} sheet(s) processed", file=sys.stderr)
        for s in ir.sheets:
            print(f"  {s.class_name} ({s.schema}, {len(s.properties)} properties)", file=sys.stderr)


# ---------------------------------------------------------------------------
# Entry point (called from cli.py)
# ---------------------------------------------------------------------------


def run(args) -> None:
    path = args.file if args.file is not None else DEFAULT_FILE
    sheet_filter = [s.strip() for s in args.sheets.split(",")] if args.sheets else None

    wb = load_workbook(path)
    ir = build_ir(
        wb, sheet_filter,
        args.namespace, args.root_class, args.dotnet_version, args.xml_docs,
        args.generate_loader, args.generate_tostring, args.generate_tojson,
        args.generate_pristine, args.readonly, args.uipath_var,
    )

    cs_output = render_cs(ir)

    if args.output_cs:
        try:
            args.output_cs.write_text(cs_output, encoding="utf-8")
        except OSError as exc:
            print(f"Error: cannot write {args.output_cs}: {exc}", file=sys.stderr)
            sys.exit(1)
    else:
        print(cs_output, end="")

    if args.output_xaml:
        xaml_output = render_xaml(ir)
        try:
            args.output_xaml.write_text(xaml_output, encoding="utf-8")
        except OSError as exc:
            print(f"Error: cannot write {args.output_xaml}: {exc}", file=sys.stderr)
            sys.exit(1)

    emit_diagnostics(ir, args.fmt)
