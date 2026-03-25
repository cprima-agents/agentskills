# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""
ConFormMold — UiPath REFramework Config.xlsx → typed C# classes and XAML clipboard snippet.

Reads ./Data/Config.xlsx (or a supplied path), detects Standard and Asset sheet
schemas, and generates:
  - A typed C# file with one class per sheet plus a CodedConfig aggregator
  - Optionally a ClipboardData XAML snippet (REF-pattern) for paste into UiPath Studio

Usage:
  uv run conform_mold.py [FILE] [OPTIONS]

Examples:
  uv run conform_mold.py
  uv run conform_mold.py ./Data/Config.xlsx --output-cs CodedConfig.cs
  uv run conform_mold.py ./Data/Config.xlsx --output-xaml LoadTypedConfig.xaml
  uv run conform_mold.py ./Data/Config.xlsx --sheets Settings,Constants
  uv run conform_mold.py ./Data/Config.xlsx --no-loader --generate-tostring
  uv run conform_mold.py ./Data/Config.xlsx --format json

Exit codes: 0 = success (warnings possible), 1 = fatal input error.
"""

import argparse
import datetime
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_FILE = Path("./Data/Config.xlsx")
DEFAULT_NAMESPACE = "Cpmf.Config"
DEFAULT_ROOT_CLASS = "CodedConfig"
DEFAULT_DOTNET = "net6"
DEFAULT_UIPATH_VAR = "out_ConFigTree"

_HEADER_STANDARD = ["name", "value", "description"]
_HEADER_ASSET = ["name", "asset", "orchestratorassetfolder", "description"]

# C# types that require `using System;`
_SYSTEM_TYPES = {"DateOnly", "DateTime", "TimeOnly"}

# ---------------------------------------------------------------------------
# Intermediate representation
# ---------------------------------------------------------------------------


@dataclass
class PropertyDef:
    prop_name: str
    cs_type: str
    description: str | None
    asset_name: str | None = None
    folder: str | None = None


@dataclass
class SheetDef:
    sheet_name: str
    class_name: str
    schema: str  # "standard" or "asset"
    properties: list[PropertyDef] = field(default_factory=list)


@dataclass
class WorkbookIR:
    root_class: str
    namespace: str
    dotnet_version: str
    emit_xml_docs: bool
    generate_loader: bool
    generate_tostring: bool
    generate_tojson: bool
    generate_pristine: bool
    readonly: bool
    uipath_var: str
    sheets: list[SheetDef] = field(default_factory=list)
    needs_orchestrator_asset: bool = False
    needs_system_using: bool = False


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------


def load_workbook(path: Path) -> openpyxl.Workbook:
    if path.suffix.lower() != ".xlsx":
        print(f"Error: expected a .xlsx file, got: {path}", file=sys.stderr)
        sys.exit(1)
    if not path.exists():
        print(f"Error: file not found: {path}", file=sys.stderr)
        sys.exit(1)
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        print(f"Error: cannot open {path}: {exc}", file=sys.stderr)
        sys.exit(1)


def detect_schema(header_row: list) -> str | None:
    """Return 'standard', 'asset', or None based on normalised header columns."""
    normalised = [str(v).strip().lower() if v is not None else "" for v in header_row]
    if len(normalised) >= 3 and normalised[:3] == _HEADER_STANDARD[:3]:
        return "standard"
    if len(normalised) >= 4 and normalised[:4] == _HEADER_ASSET:
        return "asset"
    return None


def read_sheet(ws) -> tuple[str, list[dict]]:
    """Return (schema, rows). Emits warnings and returns ('skip', []) for bad sheets."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows or all(v is None for v in all_rows[0]):
        print(f"  [WARN] Sheet '{ws.title}': no header row found, skipping.", file=sys.stderr)
        return "skip", []

    header = list(all_rows[0])
    schema = detect_schema(header)
    if schema is None:
        print(
            f"  [WARN] Sheet '{ws.title}': unrecognised header {header[:4]}, skipping.",
            file=sys.stderr,
        )
        return "skip", []

    rows = []
    for raw in all_rows[1:]:
        if not raw or raw[0] is None or str(raw[0]).strip() == "":
            continue
        if schema == "standard":
            name = str(raw[0]).strip()
            value = raw[1] if len(raw) > 1 else None
            desc = str(raw[2]).strip() if len(raw) > 2 and raw[2] is not None else None
            rows.append({"name": name, "value": value, "description": desc})
        else:  # asset
            name = str(raw[0]).strip()
            asset_name = str(raw[1]).strip() if len(raw) > 1 and raw[1] is not None else ""
            folder = str(raw[2]).strip() if len(raw) > 2 and raw[2] is not None else ""
            desc = str(raw[3]).strip() if len(raw) > 3 and raw[3] is not None else None
            rows.append({"name": name, "asset_name": asset_name, "folder": folder, "description": desc})

    return schema, rows


def infer_cs_type(py_value, dotnet_version: str) -> str:
    """Map a Python value (from openpyxl data_only) to a C# type name."""
    if py_value is None:
        return "string"
    # bool must precede int — bool is a subclass of int in Python
    if isinstance(py_value, bool):
        return "bool"
    if isinstance(py_value, int):
        return "int"
    if isinstance(py_value, float):
        return "double"
    # datetime.time before datetime.datetime
    if isinstance(py_value, datetime.time) and not isinstance(py_value, datetime.datetime):
        return "TimeOnly"
    if isinstance(py_value, datetime.datetime):
        if py_value.hour == 0 and py_value.minute == 0 and py_value.second == 0 and py_value.microsecond == 0:
            return "DateOnly"
        return "DateTime"
    if isinstance(py_value, datetime.date):
        return "DateOnly"
    return "string"


def to_pascal_case(name: str) -> str:
    """Convert a config key name to PascalCase."""
    segments = re.split(r"[^a-zA-Z0-9]+", name)
    result = "".join(seg[0].upper() + seg[1:] for seg in segments if seg)
    if result and result[0].isdigit():
        result = "_" + result
    return result


def sheet_name_to_class_name(sheet_name: str) -> str:
    """Convert sheet name to C# class name: split on . and -, PascalCase, append Config."""
    segments = re.split(r"[.\-]", sheet_name)
    result = "".join(seg[0].upper() + seg[1:] for seg in segments if seg)
    return result + "Config"


# ---------------------------------------------------------------------------
# Build IR
# ---------------------------------------------------------------------------


def build_ir(
    wb: openpyxl.Workbook,
    sheet_filter: list[str] | None,
    namespace: str,
    root_class: str,
    dotnet_version: str,
    emit_xml_docs: bool,
    generate_loader: bool,
    generate_tostring: bool,
    generate_tojson: bool,
    generate_pristine: bool,
    readonly: bool,
    uipath_var: str,
) -> WorkbookIR:
    ir = WorkbookIR(
        root_class=root_class,
        namespace=namespace,
        dotnet_version=dotnet_version,
        emit_xml_docs=emit_xml_docs,
        generate_loader=generate_loader,
        generate_tostring=generate_tostring,
        generate_tojson=generate_tojson,
        generate_pristine=generate_pristine,
        readonly=readonly,
        uipath_var=uipath_var,
    )

    requested = set(sheet_filter) if sheet_filter else None
    if requested:
        for name in requested:
            if name not in wb.sheetnames:
                print(f"  [WARN] Requested sheet '{name}' not found in workbook.", file=sys.stderr)

    for ws in wb.worksheets:
        if requested and ws.title not in requested:
            continue

        schema, rows = read_sheet(ws)
        if schema == "skip":
            continue

        class_name = sheet_name_to_class_name(ws.title)
        sheet_def = SheetDef(sheet_name=ws.title, class_name=class_name, schema=schema)

        for row in rows:
            prop_name = to_pascal_case(row["name"])
            if schema == "standard":
                cs_type = infer_cs_type(row["value"], dotnet_version)
                sheet_def.properties.append(
                    PropertyDef(prop_name=prop_name, cs_type=cs_type, description=row["description"])
                )
                if cs_type in _SYSTEM_TYPES:
                    ir.needs_system_using = True
            else:  # asset
                sheet_def.properties.append(
                    PropertyDef(
                        prop_name=prop_name,
                        cs_type="OrchestratorAsset",
                        description=row["description"],
                        asset_name=row["asset_name"],
                        folder=row["folder"],
                    )
                )

        if schema == "asset":
            ir.needs_orchestrator_asset = True

        ir.sheets.append(sheet_def)

    return ir


# ---------------------------------------------------------------------------
# C# generation
# ---------------------------------------------------------------------------


def _accessor(readonly: bool) -> str:
    return "{ get; init; }" if readonly else "{ get; set; }"


def _xml_doc(description: str | None, emit: bool, indent: str) -> str:
    if not emit or not description:
        return ""
    safe = description.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return f"{indent}/// <summary>{safe}</summary>\n"


def _property_initializer(cs_type: str) -> str:
    if cs_type == "string":
        return ' = "";'
    if cs_type == "OrchestratorAsset":
        return " = new();"
    return ""


def _from_data_table_case(prop: PropertyDef) -> list[str]:
    """Return C# switch-case lines for a single property in FromDataTable."""
    name = prop.prop_name
    lines = []
    if prop.cs_type == "OrchestratorAsset":
        lines.append(f'                case "{name}":')
        lines.append(f'                    cfg.{name}.AssetName = row[1]?.ToString()?.Trim() ?? "";')
        lines.append(f'                    cfg.{name}.Folder    = row[2]?.ToString()?.Trim() ?? "";')
        lines.append('                    break;')
    elif prop.cs_type == "string":
        lines.append(f'                case "{name}": cfg.{name} = value; break;')
    elif prop.cs_type == "int":
        lines.append(f'                case "{name}":')
        lines.append(f'                    if (int.TryParse(value, out var v_{name})) cfg.{name} = v_{name};')
        lines.append('                    break;')
    elif prop.cs_type == "double":
        lines.append(f'                case "{name}":')
        lines.append(f'                    if (double.TryParse(value, System.Globalization.NumberStyles.Any, System.Globalization.CultureInfo.InvariantCulture, out var v_{name})) cfg.{name} = v_{name};')
        lines.append('                    break;')
    elif prop.cs_type == "bool":
        lines.append(f'                case "{name}":')
        lines.append(f'                    if (bool.TryParse(value, out var v_{name})) cfg.{name} = v_{name};')
        lines.append('                    break;')
    elif prop.cs_type == "DateOnly":
        lines.append(f'                case "{name}":')
        lines.append(f'                    if (DateOnly.TryParse(value, out var v_{name})) cfg.{name} = v_{name};')
        lines.append('                    break;')
    elif prop.cs_type == "DateTime":
        lines.append(f'                case "{name}":')
        lines.append(f'                    if (DateTime.TryParse(value, out var v_{name})) cfg.{name} = v_{name};')
        lines.append('                    break;')
    elif prop.cs_type == "TimeOnly":
        lines.append(f'                case "{name}":')
        lines.append(f'                    if (TimeOnly.TryParse(value, out var v_{name})) cfg.{name} = v_{name};')
        lines.append('                    break;')
    else:
        lines.append(f'                case "{name}": cfg.{name} = value; break;')
    return lines


def render_cs_orchestrator_asset_class(readonly: bool) -> str:
    acc = _accessor(readonly)
    return (
        f"    public class OrchestratorAsset\n"
        f"    {{\n"
        f'        public string AssetName {acc} = "";\n'
        f'        public string Folder {acc} = "";\n'
        f"    }}"
    )


def render_cs_sheet_class(sheet: SheetDef, ir: WorkbookIR) -> str:
    acc = _accessor(ir.readonly)
    lines = [f"    public class {sheet.class_name}", "    {"]

    for prop in sheet.properties:
        doc = _xml_doc(prop.description, ir.emit_xml_docs, "        ")
        if doc:
            lines.append(doc.rstrip("\n"))
        init = _property_initializer(prop.cs_type)
        lines.append(f"        public {prop.cs_type} {prop.prop_name} {acc}{init}")

    if ir.generate_loader:
        lines.append("")
        lines.append(f"        public static {sheet.class_name} FromDataTable(DataTable dt)")
        lines.append("        {")
        if sheet.properties:
            lines.append(f"            var cfg = new {sheet.class_name}();")
            lines.append("            foreach (DataRow row in dt.Rows)")
            lines.append("            {")
            if sheet.schema == "standard":
                lines.append('                var key   = row[0]?.ToString()?.Trim();')
                lines.append('                var value = row[1]?.ToString()?.Trim() ?? "";')
            else:  # asset
                lines.append('                var key = row[0]?.ToString()?.Trim();')
            lines.append("                switch (key)")
            lines.append("                {")
            for prop in sheet.properties:
                lines.extend(_from_data_table_case(prop))
            lines.append("                }")
            lines.append("            }")
            lines.append("            return cfg;")
        else:
            # Empty sheet: no data rows to map
            lines.append(f"            return new {sheet.class_name}();")
        lines.append("        }")

    lines.append("    }")
    return "\n".join(lines)


def render_cs_root_class(ir: WorkbookIR) -> str:
    acc = _accessor(ir.readonly)
    lines = []
    if ir.emit_xml_docs:
        lines.append(f"    /// <summary>Root configuration object.</summary>")
    lines.append(f"    public class {ir.root_class}")
    lines.append("    {")

    for sheet in ir.sheets:
        prop_name = sheet.class_name[:-6]  # strip "Config" suffix
        lines.append(f"        public {sheet.class_name} {prop_name} {acc} = new();")

    if ir.generate_loader:
        lines.append("")
        lines.append(f"        public static {ir.root_class} Load(Dictionary<string, DataTable> tables)")
        lines.append("        {")
        lines.append(f"            var cfg = new {ir.root_class}();")
        for sheet in ir.sheets:
            prop_name = sheet.class_name[:-6]
            var_name = f"t_{prop_name}"
            lines.append(
                f'            if (tables.TryGetValue("{sheet.sheet_name}", out var {var_name}))'
                f' cfg.{prop_name} = {sheet.class_name}.FromDataTable({var_name});'
            )
        lines.append("            return cfg;")
        lines.append("        }")

    if ir.generate_pristine:
        lines.append("")
        lines.append("        public static readonly Dictionary<string, HashSet<string>> Schema = new()")
        lines.append("        {")
        for sheet in ir.sheets:
            keys = ", ".join(f'"{p.prop_name}"' for p in sheet.properties)
            lines.append(f'            ["{sheet.sheet_name}"] = new HashSet<string> {{ {keys} }},')
        lines.append("        };")
        lines.append("")
        lines.append("        public static DriftReport CheckDrift(Dictionary<string, DataTable> tables)")
        lines.append("        {")
        lines.append("            var report = new DriftReport();")
        lines.append("            foreach (var kvp in Schema)")
        lines.append("            {")
        lines.append('                if (!tables.TryGetValue(kvp.Key, out var dt)) { report.MissingSheets.Add(kvp.Key); continue; }')
        lines.append('                var actualKeys = new HashSet<string>(dt.AsEnumerable().Select(r => r[0]?.ToString()?.Trim() ?? ""));')
        lines.append('                report.UnknownKeys.AddRange(actualKeys.Except(kvp.Value).Select(k => $"{kvp.Key}.{k}"));')
        lines.append('                report.MissingKeys.AddRange(kvp.Value.Except(actualKeys).Select(k => $"{kvp.Key}.{k}"));')
        lines.append("            }")
        lines.append("            return report;")
        lines.append("        }")

    if ir.generate_tostring:
        parts = ", ".join(
            f"{sheet.class_name[:-6]} = {{{sheet.class_name[:-6]}}}"
            for sheet in ir.sheets
        )
        lines.append("")
        lines.append(f'        public override string ToString() =>')
        lines.append(f'            $"{ir.root_class} {{{{ {parts} }}}}";')

    if ir.generate_tojson:
        lines.append("")
        lines.append("        public string ToJson() =>")
        lines.append("            System.Text.Json.JsonSerializer.Serialize(this);")

    lines.append("    }")
    return "\n".join(lines)


def render_cs_drift_report_class() -> str:
    return (
        "    public class DriftReport\n"
        "    {\n"
        '        public List<string> MissingSheets { get; set; } = new();\n'
        '        public List<string> MissingKeys { get; set; } = new();\n'
        '        public List<string> UnknownKeys { get; set; } = new();\n'
        "        public bool HasDrift => MissingSheets.Any() || MissingKeys.Any() || UnknownKeys.Any();\n"
        "        public override string ToString() => HasDrift\n"
        '            ? $"Drift detected: {MissingSheets.Count} missing sheet(s), {MissingKeys.Count} missing key(s), {UnknownKeys.Count} unknown key(s)"\n'
        '            : "No drift detected";\n'
        "    }"
    )


def render_cs(ir: WorkbookIR) -> str:
    usings = []
    if ir.needs_system_using:
        usings.append("using System;")
    if ir.generate_loader or ir.generate_pristine:
        usings.append("using System.Collections.Generic;")
    if ir.generate_loader:
        usings.append("using System.Data;")
    if ir.generate_pristine:
        usings.append("using System.Linq;")
    if ir.generate_tojson:
        usings.append("using System.Text.Json;")

    sections = []
    if usings:
        sections.extend(usings)
        sections.append("")

    sections.append(f"namespace {ir.namespace}")
    sections.append("{")

    sections.append(render_cs_root_class(ir))
    sections.append("")

    if ir.generate_pristine:
        sections.append(render_cs_drift_report_class())
        sections.append("")

    if ir.needs_orchestrator_asset:
        sections.append(render_cs_orchestrator_asset_class(ir.readonly))
        sections.append("")

    for sheet in ir.sheets:
        sections.append(render_cs_sheet_class(sheet, ir))
        sections.append("")

    # Remove trailing empty line inside namespace block
    if sections and sections[-1] == "":
        sections.pop()

    sections.append("}")
    return "\n".join(sections) + "\n"


# ---------------------------------------------------------------------------
# XAML generation (REF-pattern ClipboardData clipboard snippet)
# ---------------------------------------------------------------------------


def render_xaml(ir: WorkbookIR) -> str:
    """
    Render a ClipboardData XML snippet for direct paste into UiPath Studio.

    REF-pattern:
      1. Declare dt_Tables and uipath_var variables
      2. Init dt_Tables = new Dictionary(Of String, DataTable)
      3. ExcelApplicationScope(in_ConfigFilePath) -> ForEach(in_ConfigSheets)
         -> ReadRange(in_SheetName) -> dt_Tables.Add(in_SheetName, dt_CurrentSheet)
      4. Assign: uipath_var = RootClass.Load(dt_Tables)

    Expects incoming arguments: in_ConfigFilePath (String), in_ConfigSheets (IEnumerable(Of String))
    matching the REFramework InitAllSettings interface.
    """
    var = ir.uipath_var
    root = ir.root_class

    return (
        '<?xml version="1.0" encoding="utf-16"?>\n'
        '<ClipboardData Version="1.0"\n'
        '  xmlns="http://schemas.microsoft.com/netfx/2009/xaml/activities/presentation"\n'
        '  xmlns:p="http://schemas.microsoft.com/netfx/2009/xaml/activities"\n'
        '  xmlns:sap2010="http://schemas.microsoft.com/netfx/2010/xaml/activities/presentation"\n'
        '  xmlns:scg="clr-namespace:System.Collections.Generic;assembly=System.Private.CoreLib"\n'
        '  xmlns:x="http://schemas.microsoft.com/winfx/2006/xaml"\n'
        '  xmlns:ui="http://schemas.uipath.com/workflow/activities"\n'
        '  xmlns:sd="clr-namespace:System.Data;assembly=System.Data.Common">\n'
        '  <ClipboardData.Data>\n'
        '    <scg:List x:TypeArguments="x:Object" Capacity="1">\n'
        f'      <p:Sequence x:Name="__ReferenceID0" DisplayName="{var}"\n'
        f'          sap2010:Annotation.AnnotationText="{var} typed config loader&#xD;&#xA;@see https://rpapub.github.io/ConFormMold/">\n'
        '        <p:Sequence.Variables>\n'
        '          <p:Variable x:TypeArguments="scg:Dictionary(x:String, sd:DataTable)" Name="dt_Tables" />\n'
        f'          <p:Variable x:TypeArguments="x:Object" Name="{var}" />\n'
        '        </p:Sequence.Variables>\n'
        '        <p:Assign DisplayName="Init dt_Tables">\n'
        '          <p:Assign.To>\n'
        '            <OutArgument x:TypeArguments="scg:Dictionary(x:String, sd:DataTable)">[dt_Tables]</OutArgument>\n'
        '          </p:Assign.To>\n'
        '          <p:Assign.Value>\n'
        '            <InArgument x:TypeArguments="scg:Dictionary(x:String, sd:DataTable)">[New Dictionary(Of String, DataTable)()]</InArgument>\n'
        '          </p:Assign.Value>\n'
        '        </p:Assign>\n'
        '        <ui:ExcelApplicationScope DisplayName="Open Config.xlsx" WorkbookPath="[in_ConfigFilePath]">\n'
        '          <ui:ExcelApplicationScope.Body>\n'
        '            <ActivityAction>\n'
        '              <ActivityAction.Argument>\n'
        '                <DelegateInArgument x:TypeArguments="ui:WorkbookApplication" Name="ExcelWorkbookScope" />\n'
        '              </ActivityAction.Argument>\n'
        '              <p:ForEach x:TypeArguments="x:String" DisplayName="ForEach Sheet in in_ConfigSheets" Values="[in_ConfigSheets]">\n'
        '                <p:ForEach.Body>\n'
        '                  <ActivityAction x:TypeArguments="x:String">\n'
        '                    <ActivityAction.Argument>\n'
        '                      <DelegateInArgument x:TypeArguments="x:String" Name="in_SheetName" />\n'
        '                    </ActivityAction.Argument>\n'
        '                    <p:Sequence>\n'
        '                      <p:Sequence.Variables>\n'
        '                        <p:Variable x:TypeArguments="sd:DataTable" Name="dt_CurrentSheet" />\n'
        '                      </p:Sequence.Variables>\n'
        '                      <ui:ReadRange SheetName="[in_SheetName]" AddHeaders="True" DataTable="[dt_CurrentSheet]" />\n'
        '                      <p:InvokeMethod MethodName="Add" TargetObject="[dt_Tables]">\n'
        '                        <p:InvokeMethod.Parameters>\n'
        '                          <InArgument x:TypeArguments="x:String">[in_SheetName]</InArgument>\n'
        '                          <InArgument x:TypeArguments="sd:DataTable">[dt_CurrentSheet]</InArgument>\n'
        '                        </p:InvokeMethod.Parameters>\n'
        '                      </p:InvokeMethod>\n'
        '                    </p:Sequence>\n'
        '                  </ActivityAction>\n'
        '                </p:ForEach.Body>\n'
        '              </p:ForEach>\n'
        '            </ActivityAction>\n'
        '          </ui:ExcelApplicationScope.Body>\n'
        '        </ui:ExcelApplicationScope>\n'
        '        <p:Assign DisplayName="Load typed config">\n'
        '          <p:Assign.To>\n'
        f'            <OutArgument x:TypeArguments="x:Object">[{var}]</OutArgument>\n'
        '          </p:Assign.To>\n'
        '          <p:Assign.Value>\n'
        f'            <InArgument x:TypeArguments="x:Object">[{root}.Load(dt_Tables)]</InArgument>\n'
        '          </p:Assign.Value>\n'
        '        </p:Assign>\n'
        '      </p:Sequence>\n'
        '    </scg:List>\n'
        '  </ClipboardData.Data>\n'
        '  <ClipboardData.Metadata>\n'
        '    <scg:List x:TypeArguments="x:Object" Capacity="1">\n'
        '      <x:Null />\n'
        '    </scg:List>\n'
        '  </ClipboardData.Metadata>\n'
        '</ClipboardData>\n'
    )


# ---------------------------------------------------------------------------
# Diagnostics
# ---------------------------------------------------------------------------


def emit_diagnostics(ir: WorkbookIR, fmt: str, warnings: list[str]) -> None:
    if fmt == "json":
        data = {
            "sheets": [
                {"name": s.sheet_name, "class": s.class_name, "schema": s.schema, "properties": len(s.properties)}
                for s in ir.sheets
            ],
            "needs_orchestrator_asset": ir.needs_orchestrator_asset,
            "needs_system_using": ir.needs_system_using,
            "warnings": warnings,
        }
        print(json.dumps(data, indent=2), file=sys.stderr)
    else:
        print(
            f"ConFormMold: {len(ir.sheets)} sheet(s) processed"
            + (f", {len(warnings)} warning(s)" if warnings else ""),
            file=sys.stderr,
        )
        for s in ir.sheets:
            print(f"  {s.class_name} ({s.schema}, {len(s.properties)} properties)", file=sys.stderr)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate typed C# config classes and XAML clipboard snippet from UiPath Config.xlsx."
    )
    parser.add_argument(
        "file", nargs="?", type=Path, default=None,
        metavar="FILE",
        help=f"Path to Config.xlsx (default: {DEFAULT_FILE})",
    )
    parser.add_argument("--namespace", default=DEFAULT_NAMESPACE, help="C# namespace (default: Cpmf.Config)")
    parser.add_argument(
        "--root-class", default=DEFAULT_ROOT_CLASS, dest="root_class",
        help=f"Root aggregator class name (default: {DEFAULT_ROOT_CLASS})",
    )
    parser.add_argument("--dotnet-version", default=DEFAULT_DOTNET, choices=["net6", "net8"], dest="dotnet_version")
    parser.add_argument("--no-xml-docs", action="store_false", dest="xml_docs", default=True, help="Suppress XML doc comments")
    parser.add_argument("--sheets", default=None, help="Comma-separated sheet names to include (default: all)")

    # Code-generation toggles
    parser.set_defaults(generate_loader=True)
    loader_group = parser.add_mutually_exclusive_group()
    loader_group.add_argument(
        "--generate-loader", action="store_true", dest="generate_loader",
        help="Emit Load() and FromDataTable() methods (default: on)",
    )
    loader_group.add_argument(
        "--no-loader", action="store_false", dest="generate_loader",
        help="Suppress Load() and FromDataTable() methods",
    )
    parser.add_argument(
        "--generate-tostring", action="store_true", dest="generate_tostring", default=False,
        help="Emit override string ToString() on root class",
    )
    parser.add_argument(
        "--generate-tojson", action="store_true", dest="generate_tojson", default=False,
        help="Emit string ToJson() method + using System.Text.Json",
    )
    parser.add_argument(
        "--generate-pristine", action="store_true", dest="generate_pristine", default=False,
        help="Emit Schema dict + DriftReport class for drift detection",
    )
    parser.add_argument(
        "--readonly", action="store_true", dest="readonly", default=False,
        help="Use { get; init; } instead of { get; set; }",
    )
    parser.add_argument(
        "--uipath-var", default=DEFAULT_UIPATH_VAR, dest="uipath_var",
        help=f"UiPath variable name in XAML clipboard snippet (default: {DEFAULT_UIPATH_VAR})",
    )

    parser.add_argument("--output-cs", type=Path, default=None, dest="output_cs", help="Write C# to file (default: stdout)")
    parser.add_argument("--output-xaml", type=Path, default=None, dest="output_xaml", help="Write XAML clipboard snippet to file (default: omitted)")
    parser.add_argument("--format", choices=["text", "json"], default="text", dest="fmt", help="Diagnostic output format (default: text)")
    args = parser.parse_args()

    path = args.file if args.file is not None else DEFAULT_FILE
    sheet_filter = [s.strip() for s in args.sheets.split(",")] if args.sheets else None

    wb = load_workbook(path)
    ir = build_ir(
        wb, sheet_filter,
        args.namespace, args.root_class, args.dotnet_version, args.xml_docs,
        args.generate_loader, args.generate_tostring, args.generate_tojson,
        args.generate_pristine, args.readonly, args.uipath_var,
    )

    cs_output = render_cs(ir)

    if args.output_cs:
        try:
            args.output_cs.write_text(cs_output, encoding="utf-8")
        except OSError as exc:
            print(f"Error: cannot write {args.output_cs}: {exc}", file=sys.stderr)
            sys.exit(1)
    else:
        print(cs_output, end="")

    if args.output_xaml:
        xaml_output = render_xaml(ir)
        try:
            args.output_xaml.write_text(xaml_output, encoding="utf-16")
        except OSError as exc:
            print(f"Error: cannot write {args.output_xaml}: {exc}", file=sys.stderr)
            sys.exit(1)

    emit_diagnostics(ir, args.fmt, [])


if __name__ == "__main__":
    main()
